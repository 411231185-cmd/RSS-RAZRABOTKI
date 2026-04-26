"""
Writer для PromPortal: берёт исходный XLSX как шаблон и пишет обратно тот же формат
с подменённой колонкой "Описание".

Принципы:
- Формат файла PromPortal не меняем: те же столбцы, тот же порядок, те же типы.
- Поле "Код товара" (code) не модифицируем никогда.
- Перезаписываем только колонку "Описание" из словаря {code: new_description}.
- Если для товара нет нового описания — оставляем исходное (не затираем пустотой).
"""
import logging
from pathlib import Path
from typing import Dict, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

# Имена столбцов должны совпадать с теми, что использует промпортал-ридер.
COL_CODE = "Код товара"
COL_DESCRIPTION = "Описание"


def write_promportal_xlsx(
    template_path: Path,
    output_path: Path,
    descriptions_by_code: Dict[str, str],
    sheet_name: Optional[str] = None,
) -> Dict[str, int]:
    """
    Прочитать template_path, подменить колонку "Описание" по словарю, сохранить в output_path.

    Args:
        template_path:        исходный PromPortal XLSX (структура-эталон).
        output_path:          куда сохранить результат (перезаписывается).
        descriptions_by_code: {code: новое_описание}.
        sheet_name:           имя листа. По умолчанию активный лист.

    Returns:
        {"total": N, "updated": N, "missing": N, "skipped_no_code": N}
    """
    template_path = Path(template_path)
    output_path = Path(output_path)

    if not template_path.exists():
        raise FileNotFoundError(f"Шаблон не найден: {template_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    logger.info("Открываю шаблон: %s", template_path)
    wb = load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    logger.info("Лист: '%s' | строк: %d | столбцов: %d", ws.title, ws.max_row, ws.max_column)

    # Шаг 1. Найти индексы нужных столбцов по заголовку (1-я строка).
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {str(h).strip(): idx + 1 for idx, h in enumerate(header_row) if h is not None}

    if COL_CODE not in header_map:
        raise ValueError(
            f"В шаблоне нет столбца '{COL_CODE}'. "
            f"Доступные заголовки: {list(header_map.keys())}"
        )
    if COL_DESCRIPTION not in header_map:
        raise ValueError(
            f"В шаблоне нет столбца '{COL_DESCRIPTION}'. "
            f"Доступные заголовки: {list(header_map.keys())}"
        )

    code_col = header_map[COL_CODE]
    desc_col = header_map[COL_DESCRIPTION]

    # Шаг 2. Проход по строкам данных (со 2-й строки) и подмена только колонки описания.
    total = 0
    updated = 0
    missing = 0
    skipped_no_code = 0

    for row_idx in range(2, ws.max_row + 1):
        total += 1
        code_cell = ws.cell(row=row_idx, column=code_col)
        code_value = code_cell.value

        if code_value is None or str(code_value).strip() == "":
            skipped_no_code += 1
            continue

        code = str(code_value).strip()
        new_desc = descriptions_by_code.get(code)

        if new_desc is None:
            # Для этого товара нет нового описания — НЕ трогаем существующую ячейку.
            missing += 1
            continue

        # Подменяем только колонку описания. Все остальные столбцы остаются нетронутыми.
        ws.cell(row=row_idx, column=desc_col).value = new_desc
        updated += 1

    # Шаг 3. Сохранить результат. Перезаписываем — это идемпотентно по сути операции.
    logger.info("Сохраняю результат: %s", output_path)
    wb.save(output_path)
    wb.close()

    stats = {
        "total": total,
        "updated": updated,
        "missing": missing,
        "skipped_no_code": skipped_no_code,
    }
    logger.info("Экспорт завершён: %s", stats)
    return stats